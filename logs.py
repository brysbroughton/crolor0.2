import crol
import os, sys, smtplib
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.formatting import ColorScaleRule, CellIsRule, FormulaRule

class Log(crol.GenericType):
    """
    Generic text log handling for crawl job.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'log',
            'path' : './',
            'filename' : 'croler',
            'endfilename' : '.log.txt',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'file_pointer' : None,
            "row_before" : '',
            "row_after" : '',
            "col_before" : '',
            "col_after" : '',
            "row_trim" : "right"
        }
        super(Log, self).__init__(**kwargs)
    
    def openfile(self):
        f = open(self.path+self.filename+self.endfilename, 'w')
        self.setprop('file_pointer', f)
        self.writefile(self.head_text)
    
    def writefile(self, new_val):
        self.file_pointer.write(new_val)
        self.file_pointer.flush()
        os.fsync(self.file_pointer)
    
    def closefile(self):
        self.writefile(self.foot_text)
        self.file_pointer.close
    
    def writerow(self, row):
        string_bits = []
        string_bits.append(self.row_before)
        for col in row:
            string_bits.append(self.col_before)
            string_bits.append(col)
            string_bits.append(self.col_after)
        if self.row_trim:
            if self.row_trim == 'left':
                string_bits = string_bits[1:len(string_bits)]
            elif self.row_trim == 'right':
                string_bits = string_bits[0:len(string_bits)-1]
            elif self.row_trim == 'both':
                string_bits = string_bits[1:len(string_bits)-1]
            elif self.row_trim == 'none':
                pass
            else:
                raise Exception('The only values accepted for row_trim are: "left, right, both, or none". You entered %s' % self.row_trim )
        else:
            raise Exception('Property "row_trim" must have a value')
        string_bits.append(self.row_after)
        string_bits = map(str, string_bits)
        self.writefile(''.join(string_bits))


class WebLog(Log):
    """
    HTML log handler for crawl job.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'weblog',
            'path' : './',
            'filename' : 'webLog',
            'endfilename' : '.log.html',
            'file_pointer' : None,
            'html_before' : '<!DOCTYPE html><html>\n<head></head>\n<body>',
            'html_after' : '\n</body></html>',
            'table_wrapper' : '\n<div class="table">',
            'row_wrapper' : '\n\n<div class="row">',
            'col_wrapper' : '\n<div class="col">',
            'wrapper_after' : '</div>',
            'default_headings' : [],
            'crawl_report' : None
        }
        crol.GenericType.__init__(self, **kwargs)
        self.injectcss()
        
        if self.crawl_report:
            self.readreport(self.crawl_report)
    
    def openfile(self):
        """
        Opens file based on endfilename.
        Assigns the open file to file_pointer.
        """
        
        f = open(self.path+self.filename+self.endfilename, 'w')
        self.setprop('file_pointer', f)
    
    def reporttofile(self, crawl_report):
        """
        Reads the given CrawlReport.
        Writes the url_data and report_stats into the opened file.
        """
        
        self.setprop('crawl_report', crawl_report)
        self.openfile()
        self.writefile(self.html_before)
        
        #write in report_stats from crawl_report
        self.writefile(self.table_wrapper)
        self.writerow(['STATISTIC', 'VALUE'])
        self.writerow(['Total urls found:', self.crawl_report.statistics['total_count']])
        self.writerow(['OK urls found:', self.crawl_report.statistics['ok_count']])
        self.writerow(['Redirected urls found:', self.crawl_report.statistics['redirected_count']])
        self.writerow(['Broken urls found:', self.crawl_report.statistics['broken_count']])
        self.writefile(self.wrapper_after)
        
        #write in url_data from crawl_report
        self.writefile(self.table_wrapper)
        self.writerow(['STATUS', 'REASON', 'MIMETYPE', 'URL', 'PARENT'])
        for report in self.crawl_report.url_reports:
            self.writerow([report.status, report.reason, report.mimetype, report.url, report.parent_url])
        self.writefile(self.wrapper_after)
        
        self.writefile(self.html_after)
        self.closefile()
    
    def closefile(self):
        """
        Closes the currently open file.
        """
        
        self.file_pointer.close
    
    def writerow(self, row):
        """
        Builds a string of HTML code from the given row.
        Returns the HTML code as a string.
        """
        
        string_bits = []
        string_bits.append(self.row_wrapper)
        # Build the first col as a colored div
        string_bits.append(self.col_wrapper.replace('class="col"', 'class="col %s"' % self.statuscolor(row[0])))
        string_bits.append(str(row[0]).replace('<','&lt;').replace('>','&gt;'))
        string_bits.append(self.wrapper_after)
        # Build the rest of the row normally
        for col in row[1:]:
            if self.isurl(col): string_bits.append(self.col_wrapper + '<a href="'+col+'" target="_blank">')
            else: string_bits.append(self.col_wrapper)
            string_bits.append(str(col).replace('<','&lt;').replace('>','&gt;'))
            if self.isurl(col): string_bits.append('</a>' + self.wrapper_after)
            else: string_bits.append(self.wrapper_after)
        string_bits.append(self.wrapper_after)
        string_bits = map(str, string_bits)
        self.writefile(''.join(string_bits))
    
    def isurl(self, string):
        """
        Returns boolean based on the given url being http:// or https://.
        """
        
        string = str(string)
        if string.startswith('http://') or string.startswith('https://'): return True
        else: return False
    
    def statuscolor(self, status):
        """
        Returns the corresponding color based on the given status code.
        """
        
        #1xx informational status
        if isinstance(status, (int, long)) and str(status).startswith('1'): return 'blue'
        #2xx success status
        if isinstance(status, (int, long)) and str(status).startswith('2'): return 'green'
        #3xx redirection status
        if isinstance(status, (int, long)) and str(status).startswith('3'): return 'orange'
        #4xx client error status
        if isinstance(status, (int, long)) and str(status).startswith('4'): return 'red'
        #5xx client server error
        if isinstance(status, (int, long)) and str(status).startswith('5'): return 'purple'
        #no http status
        else: return None
    
    def injectcss(self):
        """
        Reads the css from weblog.css and writes it directly into html_before.
        """
        
        try:
            css_file = open('weblog.css', 'r')
            css = css_file.read()
            self.html_before = self.html_before.replace('<head></head>', '<head><style type="text/css">\n'+css+'\n</style></head>')
        except IOError as error:
            print 'Error opening weblog.css file.'


class CsvLog(Log):
    """
    CSV log handler for crawl job.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'log',
            'path' : './',
            'filename' : 'csvLog',
            'endfilename' : '.log.csv',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'file_pointer' : None,
            "row_before" : '',
            "row_after" : '\n',
            "col_before" : '',
            "col_after" : ',',
            "heading_row" : [],
            "row_trim" : "right"
        }
        crol.GenericType.__init__(self, **kwargs)

    def headingrow(self, headings=None):
        if headings:
            self.writefile(self.row_before)
            for header in  headings:
                self.writefile(self.col_before)
                self.writefile(header)
                self.writefile(self.col_after)
        else:
            self.writefile(self.heading_row)
        self.writefile(self.row_after)

class ExcelLog(Log):
    """
    Excel log handler for crawl job.
    """
    
    def __init__(self, kwargs={}):
        self.props = {
            'type' : 'excelLog',
            'path' : './',
            'filename' : 'excelLog',
            'endfilename' : '.xlsx',
            'head_text' : 'header\n',
            'foot_text' : 'footer',
            'filePointer' : None,
            "row_before" : '',
            "row_after" : '',
            "col_before" : '',
            "col_after" : '',
            "heading_row" : [],
            "row_trim" : "none",
            "workbook" : None,
            "worksheet" : None,
            "numrows" : 1,
            'crawl_report' : None
        }
        crol.GenericType.__init__(self, **kwargs)

    
    def openfile(self):
        wb = Workbook()
        ws = wb.active
        self.workbook = wb
        self.worksheet = ws
        #self.headingrow(['STATUS', 'REASON', 'MIMETYPE', 'URL', 'PARENT'])
    
    def closefile(self):
        self.workbook.save(self.path+self.filename + self.endfilename)
    
    def writerow(self, vals):
        ws = self.workbook.active
        this_row = self.numrows
        this_col = 1
        #code_color = None
        code_color = PatternFill(start_color='00FFCC00', fill_type='solid')
        index = 1
        if str(vals)[1] == "'":
            index = 2
        if str(vals)[index] == '1':
            code_color = PatternFill(start_color='004472B9', end_color='004472B9', fill_type='solid') #blue
        elif str(vals)[index] == '2':
            code_color = PatternFill(start_color='004CA454', end_color='004CA454', fill_type='solid') #green
        elif str(vals)[index] == '3':
            code_color = PatternFill(start_color='00D49b00', end_color='00D49b00', fill_type='solid') #orange
        elif str(vals)[index] == '4':
            code_color = PatternFill(start_color='00BE4C39', end_color='00BE4C39', fill_type='solid') #red
        elif str(vals)[index] == '5':
            code_color = PatternFill(start_color='009351A6', end_color='009351A6', fill_type='solid') #purple
            
        for v in vals:
            cell = ws.cell(row = this_row, column = this_col)
            cell.value = v
            cell.fill = code_color
            if ws.column_dimensions[get_column_letter(this_col)].width < len(str(v)):
                ws.column_dimensions[get_column_letter(this_col)].width = len(str(v)) + 4
            this_col += 1
        self.numrows += 1
        self.worksheet = ws
    
    def headingrow(self, headings=None):
        ws = self.workbook.active
        if headings:
            this_col = 1
            this_row = self.numrows
            for header in  headings:
                cell = ws.cell(row = this_row, column = this_col)
                cell.value = header
                cell.font = Font(size=16)
                if ws.column_dimensions[get_column_letter(this_col)].width < len(str(header)):
                    ws.column_dimensions[get_column_letter(this_col)].width = len(str(header)) + 4
                this_col += 1
        else:
            this_col = 1
            this_row = self.numrows
            for header in  self.heading_row:
                cell = ws.cell(row = this_row, column = this_col)
                cell.value = header
                cell.font = Font(size=16)
                if ws.column_dimensions[get_column_letter(this_col)].width < len(str(header)):
                    ws.column_dimensions[get_column_letter(this_col)].width = len(str(header)) + 4
                this_col += 1
        self.numrows += 1
        self.worksheet = ws
        
    def reporttofile(self, crawl_report):
        """
        Reads the given CrawlReport.
        Writes the url_data and report_stats into the opened file.
        """
        
        self.setprop('crawl_report', crawl_report)
        self.openfile()
        self.writerow(['STATISTIC', 'VALUE'])
        self.writerow(['Total urls found:', self.crawl_report.statistics['total_count']])
        self.writerow(['OK urls found:', self.crawl_report.statistics['ok_count']])
        self.writerow(['Redirected urls found:', self.crawl_report.statistics['redirected_count']])
        self.writerow(['Broken urls found:', self.crawl_report.statistics['broken_count']])
        
        #write in url_data from crawl_report
        self.headingrow(['STATUS', 'REASON', 'MIMETYPE', 'URL', 'PARENT'])
        for report in self.crawl_report.url_reports:
            self.writerow([report.status, report.reason, report.mimetype, report.url, report.parent_url])
        
        self.closefile()
